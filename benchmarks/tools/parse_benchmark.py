import re
import pandas as pd
import argparse
from pathlib import Path
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment


def extract_benchmark_blocks(log_text):
    """Extract all benchmark result blocks from the log"""
    pattern = r"(Traffic request rate:[\s\S]*?==================================================)"
    return re.findall(pattern, log_text)


def parse_benchmark_block(block):
    """Parse a single benchmark result block"""
    data = {}

    # Extract basic information
    base_info_patterns = {
        "Req Rate": r"Traffic request rate:\s+([\d.]+)",
        "Burstiness": r"Burstiness factor:\s+([\d.]+)",
        "Max Concurrency": r"Maximum request concurrency:\s+(\d+)"
    }

    for field, pattern in base_info_patterns.items():
        match = re.search(pattern, block)
        data[field] = float(match.group(1)) if match and '.' in match.group(1) else int(
            match.group(1)) if match else None

    # Extract result metrics
    result_patterns = {
        "Success Req": r"Successful requests:\s+(\d+)",
        "Duration(s)": r"Benchmark duration \(s\):\s+([\d.]+)",
        "Input Tokens": r"Total input tokens:\s+(\d+)",
        "Output Tokens": r"Total generated tokens:\s+(\d+)",
        "Req Throughput": r"Request throughput \(req/s\):\s+([\d.]+)",
        "Out Throughput": r"Output token throughput \(tok/s\):\s+([\d.]+)",
        "Total Throughput": r"Total Token throughput \(tok/s\):\s+([\d.]+)",
        "TTFT_Mean": r"Mean TTFT \(ms\):\s+([\d.]+)",
        "TTFT_Median": r"Median TTFT \(ms\):\s+([\d.]+)",
        "TTFT_P90": r"P90 TTFT \(ms\):\s+([\d.]+)",
        "TTFT_P95": r"P95 TTFT \(ms\):\s+([\d.]+)",
        "TTFT_P99": r"P99 TTFT \(ms\):\s+([\d.]+)",
        "TPOT_Mean": r"Mean TPOT \(ms\):\s+([\d.]+)",
        "TPOT_Median": r"Median TPOT \(ms\):\s+([\d.]+)",
        "TPOT_P90": r"P90 TPOT \(ms\):\s+([\d.]+)",
        "TPOT_P95": r"P95 TPOT \(ms\):\s+([\d.]+)",
        "TPOT_P99": r"P99 TPOT \(ms\):\s+([\d.]+)",
        "ITL_Mean": r"Mean ITL \(ms\):\s+([\d.]+)",
        "ITL_Median": r"Median ITL \(ms\):\s+([\d.]+)",
        "ITL_P90": r"P90 ITL \(ms\):\s+([\d.]+)",
        "ITL_P95": r"P95 ITL \(ms\):\s+([\d.]+)",
        "ITL_P99": r"P99 ITL \(ms\):\s+([\d.]+)",
        "E2E_Mean": r"Mean E2EL \(ms\):\s+([\d.]+)",
        "E2E_Median": r"Median E2EL \(ms\):\s+([\d.]+)",
        "E2E_P90": r"P90 E2EL \(ms\):\s+([\d.]+)",
        "E2E_P95": r"P95 E2EL \(ms\):\s+([\d.]+)",
        "E2E_P99": r"P99 E2EL \(ms\):\s+([\d.]+)"
    }

    for field, pattern in result_patterns.items():
        match = re.search(pattern, block)
        data[field] = float(match.group(1)) if match and '.' in match.group(1) else int(
            match.group(1)) if match else None

    return data


def create_excel_with_merged_cells(results, output_path):
    """Create an Excel file with merged header cells"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"

    # Create a DataFrame
    df = pd.DataFrame(results)

    # Define metric groups and display names
    metric_groups = [
        {"name": "TTFT", "title": "TTFT (ms)", "metrics": ["Mean", "Median", "P90", "P95", "P99"]},
        {"name": "TPOT", "title": "TPOT (ms)", "metrics": ["Mean", "Median", "P90", "P95", "P99"]},
        {"name": "ITL", "title": "ITL (ms)", "metrics": ["Mean", "Median", "P90", "P95", "P99"]},
        {"name": "E2E", "title": "E2E (ms)", "metrics": ["Mean", "Median", "P90", "P95", "P99"]}
    ]

    # Build header rows
    headers = [
        'Req Rate', 'Burstiness', 'Max Concurrency', 'Success Req', 'Duration(s)',
        'Input Tokens', 'Output Tokens', 'Req Throughput', 'Out Throughput', 'Total Throughput'
    ]

    # Add group headers with empty cells for merging
    for group in metric_groups:
        headers.append(group["title"])
        headers.extend([""] * (len(group["metrics"]) - 1))

    # Add second header row (individual metrics)
    second_header = [""] * 10  # First 10 columns for base metrics
    for group in metric_groups:
        second_header.extend(group["metrics"])

    # Write headers
    ws.append(headers)
    ws.append(second_header)

    # Merge cells for metric group headers
    col = 11  # Metric groups start from column 11
    for group in metric_groups:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(group["metrics"]) - 1)
        ws.cell(row=1, column=col, value=group["title"])
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        col += len(group["metrics"])

    # Prepare and write data rows
    for _, row in df.iterrows():
        data_row = [
            row['Req Rate'], row['Burstiness'], row['Max Concurrency'],
            row['Success Req'], row['Duration(s)'], row['Input Tokens'],
            row['Output Tokens'], row['Req Throughput'], row['Out Throughput'],
            row['Total Throughput']
        ]

        for group in metric_groups:
            for metric in group["metrics"]:
                data_row.append(row[f"{group['name']}_{metric}"])

        ws.append(data_row)

    # Save file
    wb.save(output_path)
    return df


def main():
    parser = argparse.ArgumentParser(description='Parse benchmark log and generate an Excel report')
    parser.add_argument('log_file', help='Path to the input log file')
    parser.add_argument('-o', '--output', help='Output Excel file path (default: same as log file with .xlsx)')
    args = parser.parse_args()

    log_path = Path(args.log_file)
    if not log_path.exists():
        print(f"Error: Log file not found - {log_path}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output) if args.output else log_path.with_suffix('.xlsx')

    with open(log_path, 'r') as f:
        log_content = f.read()

    benchmark_blocks = extract_benchmark_blocks(log_content)
    if not benchmark_blocks:
        print(f"Warning: No valid benchmark results found in {log_path}", file=sys.stderr)
        sys.exit(0)

    all_results = []
    for i, block in enumerate(benchmark_blocks, 1):
        try:
            result = parse_benchmark_block(block)
            all_results.append(result)
            print(f"✓ Successfully parsed benchmark #{i}")
        except Exception as e:
            print(f"× Failed to parse benchmark #{i}: {str(e)}", file=sys.stderr)

    if not all_results:
        print("Error: No valid results were parsed", file=sys.stderr)
        sys.exit(1)

    try:
        df = create_excel_with_merged_cells(all_results, output_path)
        print(f"\n✓ Successfully saved {len(df)} records to {output_path}")
        print("\nPreview of data:")
        print(df.to_string(index=False))

    except Exception as e:
        print(f"\n× Failed to save Excel file: {str(e)}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()